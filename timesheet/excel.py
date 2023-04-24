import pandas as pd

from timesheet.configer import get_config

def format_excel(new_file, sheet_name):
    # Format excel

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Color
    #pattern = PatternFill(bgColor="#346A8C", fill_type = "solid")
    wb = load_workbook(filename = new_file)
    wb.active = wb[sheet_name]
    ws = wb.active


    blueFill = PatternFill(start_color='00CCFFFF',
                    end_color='00CCFFFF',
                    fill_type='solid')
    orangeFill = PatternFill(start_color='00FFCC99',
                    end_color='00FFCC99',
                    fill_type='solid')
    pinkFill = PatternFill(start_color='00FF99CC',
                    end_color='00FF99CC',
                    fill_type='solid')

    def format_cell(row_number, style):
        ws.cell(row_number, 2).fill = style
        ws.cell(row_number, 3).fill = style
        ws.cell(row_number, 4).fill = style


    config = get_config()
    if config:
        for row_number in range(1,ws.max_row+1):
            for dictonary in config["colors"]:
                for key in dictonary.keys():
                    if ws.cell(row_number, 2).value == key:
                        format_cell(row_number, PatternFill(start_color=Color(dictonary[key]),
                                                            end_color=Color(dictonary[key]),
                                                            fill_type='solid'))
    #resize
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    

    for col, value in dims.items():
        ws.column_dimensions[str(col)].width = value

    wb.save(new_file)


def read_excel(filename, drop=True):
    df = pd.read_excel(filename, sheet_name="Timesheet")
    if drop:
        df.drop('Startzeit', axis='columns', inplace=True)
        df.drop('Endzeit', axis='columns', inplace=True)
        df.drop('Nummer', axis='columns', inplace=True)
        df.drop('Dauer', axis='columns', inplace=True)
        df.drop('Beschreibung', axis='columns', inplace=True)
    df = df.rename(columns={"Dauer (rel.)": "Arbeitszeit"})
    return df
