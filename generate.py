import numpy as np
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import matplotlib.pyplot as plt
from datetime import datetime
from pathlib import Path 
import yaml
import glob
import os
import math
import plotext as plt
import math

def round_hours(hours):
    full_hours = math.floor(hours)
    part = hours % 1
    if part == 0 or part == 0.25 or part == 0.5 or part == 0.75:
        return hours
    if part < 1 and part > 0.75:
        full_hours = full_hours + 1 
        part = 0
    if part < 0.75 and part > 0.5:
        part = 0.75
    if part < 0.5 and part > 0.25:
        part = 0.5
    if part > 0 and part < 0.25:        
        part = 0.25
    return full_hours+part

def csv_export(df):
    fileName, elements, round = None,{}, False
    with open('config.yaml') as f:
        config = yaml.load(f, Loader=yaml.FullLoader)
        try: 
            fileName = config["csv"]["fileName"]
            elements = config["csv"]["elements"]
            round = config["csv"]["round"]
        except:
            return None

    df2 = df[df.Projekt.isin(elements.keys())]
    df2["element"] = df2["Projekt"].map(elements)
    
    df2 = df2.rename(columns={"Datum": "date"})
    df2 = df2.rename(columns={"Arbeitszeit": "hours"})
    df2.sort_values(by='date', inplace=True)
    #df2['date'] = df2['date'].dt.strftime('%d.%m.%Y')

    df_grp = df2.sort_values(['date','element'],ascending=False).groupby(['date','element']).sum()
    if round: 
        df_grp["hours"] = df_grp["hours"].map(round_hours)
    df_grp["comment"] = ""
    df_grp["aktivitätencode"] = ""
    #df_grp['date'] = df_grp['date'].dt.strftime('%d.%m.%Y')
    df_grp.to_csv(fileName,sep=';')

    fd = pd.read_csv(fileName, sep=";")
    fd["date"] = pd.to_datetime(fd["date"])
    fd['date'] = fd['date'].dt.strftime('%d.%m.%Y')
    fd.to_csv(fileName,sep=';', columns=["date","element","hours","comment","aktivitätencode"], header=True, index=False)
    pass

def format_excel(new_file, sheet_name):
    # Format excel

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Color
    #pattern = PatternFill(bgColor="#346A8C", fill_type = "solid")
    wb = load_workbook(filename = new_file)
    wb.active = wb[sheet_name]
    ws = wb.active


    blueFill = PatternFill(start_color='00CCFFFF',
                    end_color='00CCFFFF',
                    fill_type='solid')
    orangeFill = PatternFill(start_color='00FFCC99',
                    end_color='00FFCC99',
                    fill_type='solid')
    pinkFill = PatternFill(start_color='00FF99CC',
                    end_color='00FF99CC',
                    fill_type='solid')

    def format_cell(row_number, style):
        ws.cell(row_number, 2).fill = style
        ws.cell(row_number, 3).fill = style
        ws.cell(row_number, 4).fill = style

    with open('config.yaml') as f:
        config = yaml.load(f, Loader=yaml.FullLoader)
        for row_number in range(1,ws.max_row+1):
            for dictonary in config["colors"]:
                for key in dictonary.keys():
                    if ws.cell(row_number, 2).value == key:
                        format_cell(row_number, PatternFill(start_color=Color(dictonary[key]),
                                                            end_color=Color(dictonary[key]),
                                                            fill_type='solid'))
    #resize
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    

    for col, value in dims.items():
        ws.column_dimensions[str(col)].width = value

    wb.save(new_file)

def print_plotext(items, values):
    plt.bar(items, values, orientation = 'h', width = 3/5)
    plt.title('Verteilung')
    plt.clc()
    plt.plotsize(100, 2 * len(items) + 4)
    plt.show()
    
def visual(percent):
    counter = math.floor(float(percent)/3)
    rval = ""
    for i in range (counter):
        rval += "#"
    return rval

def read_excel(filename):
    df = pd.read_excel(filename, sheet_name="Timesheet")
    df.drop('Startzeit',axis='columns', inplace=True)
    df.drop('Endzeit',axis='columns', inplace=True)
    df.drop('Nummer',axis='columns', inplace=True)
    df.drop('Dauer',axis='columns', inplace=True)
    df.drop('Beschreibung',axis='columns', inplace=True)
    df = df.rename(columns={"Dauer (rel.)": "Arbeitszeit"})
    return df

def main():
    list_of_files = glob.glob('./T*.xls') # * means all if need specific format then *.csv
    latest_file = Path(max(list_of_files, key=os.path.getctime))
    print(f"Benutzte Daten: {latest_file}")

    df = read_excel(latest_file)

    df_excel = df.groupby(['Datum','Projekt']).sum()
    sheet_name = "Timmesheet"
    new_file = "./Calced_timesheet.xlsx"
    with pd.ExcelWriter(new_file, mode='w', engine='openpyxl') as writer:  
        df_excel.to_excel(writer, sheet_name= sheet_name)

    csv_export(df)

    #Information for cli
    print("\n#################\nWochen:\n")

    df['Woche'] = df['Datum'].dt.isocalendar().week

    df_week = df.groupby(['Woche']).agg({"Arbeitszeit": np.sum, "Datum": pd.Series.nunique})
    df_week['Soll_Zeit']= df_week['Datum'] * 7.8
    df_week['Diff'] = df_week['Arbeitszeit'] - df_week['Soll_Zeit']
    print(df_week)
    with pd.ExcelWriter(new_file, mode='a', engine='openpyxl') as writer:  
        df_week.to_excel(writer, sheet_name= "Zeit pro Woche")
    print(f"\n\tSumme der Differenz: {'{0:.2f}h'.format(df_week['Diff'].sum())}")



    df_prj = df.groupby(['Projekt']).sum()
    df_prj['percent']  = (df_prj['Arbeitszeit'] / df_prj['Arbeitszeit'].sum()) * 100
    df_prj['percent_visual'] = df_prj.apply(lambda row : visual(row['percent']), axis = 1)
    df_prj['percent'] = pd.Series(["{0:.2f}%".format(val) for val in df_prj['percent']], index = df_prj.index)


    df_perWeek = df.groupby(['Woche','Projekt']).agg({'Arbeitszeit':['sum']})
    df_perWeek['percent'] = df_perWeek.Arbeitszeit['sum']/39 *100
    df_perWeek['percent'] = pd.Series(["{0:.0f}%".format(val) for val in df_perWeek['percent']], index = df_perWeek.index)

    sheet_name2 = "ProjectPerWeek"
    with pd.ExcelWriter(new_file, mode='a', engine='openpyxl') as writer:  
        df_perWeek.to_excel(writer, sheet_name= sheet_name2)

    df_prj.drop('Woche',axis='columns', inplace=True)
    print("\nVerteilung:")
    print(df_prj)
    print_plotext(df_prj.index.array.to_numpy(),df_prj['Arbeitszeit'].to_numpy())

    print(f"Geschrieben: {new_file}")

    format_excel(new_file, sheet_name2)
    format_excel(new_file, sheet_name)

if __name__ == '__main__':
    main()
   